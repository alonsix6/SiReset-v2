"""
Generador de Excel Mougli

Genera archivo Excel con 1-3 hojas según datos disponibles:
- Solo Monitor → 1 hoja (Monitor)
- Solo OutView → 1 hoja (OutView)
- Ambos → 3 hojas (Monitor, OutView, Consolidado)

Cada hoja incluye:
- Filas 1-8: Metadatos
- Fila 9: Headers
- Fila 10+: Datos
"""

import logging
import io
import pandas as pd
from typing import Optional
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from app.processors.consolidador import (
    consolidar_monitor_outview,
    crear_metadatos_consolidado
)

logger = logging.getLogger('mougli.excel_generator')


# ==========================================
# FUNCIÓN PRINCIPAL
# ==========================================

def generar_excel_mougli_completo(
    df_monitor: Optional[pd.DataFrame] = None,
    df_outview: Optional[pd.DataFrame] = None,
    metadatos_monitor: Optional[pd.DataFrame] = None,
    metadatos_outview: Optional[pd.DataFrame] = None
) -> io.BytesIO:
    """
    Genera Excel completo con 1-3 hojas según datos disponibles

    Args:
        df_monitor: DataFrame Monitor procesado (39 columnas)
        df_outview: DataFrame OutView procesado (33 columnas)
        metadatos_monitor: Metadatos Monitor (opcional, se genera si no se provee)
        metadatos_outview: Metadatos OutView (opcional, se genera si no se provee)

    Returns:
        BytesIO con archivo Excel SiReset_Mougli.xlsx

    Raises:
        ValueError: Si ambos DataFrames son None
    """

    if df_monitor is None and df_outview is None:
        raise ValueError("Debe proveer al menos Monitor o OutView")

    logger.info("Generando Excel Mougli completo...")

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # ==========================================
        # HOJA 1: Monitor (si existe)
        # ==========================================

        if df_monitor is not None:
            logger.info(f"Escribiendo hoja Monitor ({len(df_monitor)} filas)")

            # Generar metadatos si no se proveyeron
            if metadatos_monitor is None:
                metadatos_monitor = _crear_metadatos_monitor(df_monitor)

            # Escribir metadatos (filas 1-8)
            metadatos_monitor.to_excel(
                writer,
                sheet_name='Monitor',
                startrow=0,
                index=False,
                header=False
            )

            # Escribir datos (fila 9+)
            df_monitor.to_excel(
                writer,
                sheet_name='Monitor',
                startrow=8,
                index=False
            )

            # Aplicar formato
            worksheet = writer.sheets['Monitor']
            aplicar_formato_monitor(worksheet, df_monitor)

        # ==========================================
        # HOJA 2: OutView (si existe)
        # ==========================================

        if df_outview is not None:
            logger.info(f"Escribiendo hoja OutView ({len(df_outview)} filas)")

            # Generar metadatos si no se proveyeron
            if metadatos_outview is None:
                metadatos_outview = _crear_metadatos_outview(df_outview)

            # Escribir metadatos (filas 1-8)
            metadatos_outview.to_excel(
                writer,
                sheet_name='OutView',
                startrow=0,
                index=False,
                header=False
            )

            # Escribir datos (fila 9+)
            df_outview.to_excel(
                writer,
                sheet_name='OutView',
                startrow=8,
                index=False
            )

            # Aplicar formato
            worksheet = writer.sheets['OutView']
            aplicar_formato_outview(worksheet, df_outview)

        # ==========================================
        # HOJA 3: Consolidado (solo si ambos existen)
        # ==========================================

        if df_monitor is not None and df_outview is not None:
            logger.info("Escribiendo hoja Consolidado")

            # Consolidar datos
            df_consolidado = consolidar_monitor_outview(df_monitor, df_outview)
            logger.info(f"Consolidado: {len(df_consolidado)} filas")

            # Generar metadatos
            metadatos_consolidado = crear_metadatos_consolidado(df_consolidado)

            # Escribir metadatos (filas 1-8)
            metadatos_consolidado.to_excel(
                writer,
                sheet_name='Consolidado',
                startrow=0,
                index=False,
                header=False
            )

            # Escribir datos (fila 9+)
            df_consolidado.to_excel(
                writer,
                sheet_name='Consolidado',
                startrow=8,
                index=False
            )

            # Aplicar formato
            worksheet = writer.sheets['Consolidado']
            aplicar_formato_consolidado(worksheet, df_consolidado)

    output.seek(0)

    # Log resumen
    hojas_creadas = []
    if df_monitor is not None:
        hojas_creadas.append('Monitor')
    if df_outview is not None:
        hojas_creadas.append('OutView')
    if df_monitor is not None and df_outview is not None:
        hojas_creadas.append('Consolidado')

    logger.info(f"Excel generado exitosamente con {len(hojas_creadas)} hoja(s): {', '.join(hojas_creadas)}")

    return output


# ==========================================
# FUNCIONES AUXILIARES DE METADATOS
# ==========================================

def _crear_metadatos_monitor(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea DataFrame de metadatos para hoja Monitor

    Returns:
        DataFrame con 8 filas de metadatos
    """

    filas = len(df)

    # Rango de fechas
    if 'DIA' in df.columns:
        fecha_min = pd.to_datetime(df['DIA']).min().strftime('%d/%m/%Y')
        fecha_max = pd.to_datetime(df['DIA']).max().strftime('%d/%m/%Y')
        rango_fechas = f"{fecha_min} - {fecha_max}"
    else:
        rango_fechas = "N/A"

    # Marcas únicas
    marcas_unicas = df['MARCA'].nunique() if 'MARCA' in df.columns else 0

    # Sectores únicos
    if 'SECTOR' in df.columns:
        sectores_unicos = sorted(df['SECTOR'].dropna().unique())
        sectores = ', '.join(sectores_unicos[:3])
        if len(sectores_unicos) > 3:
            sectores += '...'
    else:
        sectores = ""

    # Categorías únicas
    if 'CATEGORIA' in df.columns:
        categorias_unicas = sorted(df['CATEGORIA'].dropna().unique())
        categorias = ', '.join(categorias_unicas[:5])
        if len(categorias_unicas) > 5:
            categorias += '...'
    else:
        categorias = ""

    # Regiones únicas
    if 'REGION/ÁMBITO' in df.columns:
        regiones = ', '.join(sorted(df['REGION/ÁMBITO'].dropna().unique()))
    else:
        regiones = ""

    # Crear DataFrame de metadatos
    metadatos_data = [
        ['Descripción', 'Valor'],
        ['Filas', filas],
        ['Rango de fechas', rango_fechas],
        ['Marcas / Anunciantes', marcas_unicas],
        ['Sectores', sectores],
        ['Categorías', categorias],
        ['Regiones', regiones],
        ['', '']  # Fila vacía
    ]

    return pd.DataFrame(metadatos_data)


def _crear_metadatos_outview(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea DataFrame de metadatos para hoja OutView

    Returns:
        DataFrame con 8 filas de metadatos
    """

    filas = len(df)

    # Rango de fechas
    if 'Fecha' in df.columns:
        fecha_min = pd.to_datetime(df['Fecha']).min().strftime('%d/%m/%Y')
        fecha_max = pd.to_datetime(df['Fecha']).max().strftime('%d/%m/%Y')
        rango_fechas = f"{fecha_min} - {fecha_max}"
    else:
        rango_fechas = "N/A"

    # Marcas únicas
    marcas_unicas = df['Marca'].nunique() if 'Marca' in df.columns else 0

    # Sectores únicos
    if 'Sector' in df.columns:
        sectores_unicos = sorted(df['Sector'].dropna().unique())
        sectores = ', '.join(sectores_unicos[:3])
        if len(sectores_unicos) > 3:
            sectores += '...'
    else:
        sectores = ""

    # Categorías únicas
    if 'Categoría' in df.columns:
        categorias_unicas = sorted(df['Categoría'].dropna().unique())
        categorias = ', '.join(categorias_unicas[:5])
        if len(categorias_unicas) > 5:
            categorias += '...'
    else:
        categorias = ""

    # Regiones únicas
    if 'Región' in df.columns:
        regiones = ', '.join(sorted(df['Región'].dropna().unique()))
    else:
        regiones = ""

    # Crear DataFrame de metadatos
    metadatos_data = [
        ['Descripción', 'Valor'],
        ['Filas', filas],
        ['Rango de fechas', rango_fechas],
        ['Marcas / Anunciantes', marcas_unicas],
        ['Sectores', sectores],
        ['Categorías', categorias],
        ['Regiones', regiones],
        ['', '']  # Fila vacía
    ]

    return pd.DataFrame(metadatos_data)


# ==========================================
# FUNCIONES DE FORMATO - MONITOR
# ==========================================

def aplicar_formato_monitor(worksheet, df: pd.DataFrame):
    """
    Aplica formato a hoja Monitor
    """

    # 1. Formato metadatos (fila 2)
    header_fill = PatternFill(
        start_color='D9E1F2',
        end_color='D9E1F2',
        fill_type='solid'
    )
    bold_font = Font(bold=True)

    for col in range(1, 3):  # Columnas A-B
        worksheet.cell(row=2, column=col).font = bold_font
        worksheet.cell(row=2, column=col).fill = header_fill

    # 2. Formato headers (fila 9)
    header_fill_datos = PatternFill(
        start_color='4472C4',
        end_color='4472C4',
        fill_type='solid'
    )
    header_font = Font(bold=True, color='FFFFFF')

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=9, column=col)
        cell.fill = header_fill_datos
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # 3. Formato fechas (columna DIA)
    for row in range(10, len(df) + 10):
        worksheet.cell(row=row, column=1).number_format = 'DD/MM/YYYY'

    # 4. Formato INVERSIÓN (columna según posición)
    # La columna INVERSION está en la posición que corresponde
    cols_dict = {col: idx + 1 for idx, col in enumerate(df.columns)}
    if 'INVERSION' in cols_dict:
        col_inv = cols_dict['INVERSION']
        for row in range(10, len(df) + 10):
            worksheet.cell(row=row, column=col_inv).number_format = '#,##0.00'

    # 5. Ajustar anchos de columna
    ajustar_anchos_monitor(worksheet)


def ajustar_anchos_monitor(worksheet):
    """
    Ajusta anchos de columnas para Monitor
    """
    ANCHOS = {
        'A': 12,   # DIA
        'B': 6,    # AÑO
        'C': 12,   # MES
        'D': 8,    # SEMANA
        'E': 15,   # MEDIO
        'F': 15,   # MARCA
        'G': 20,   # PRODUCTO
        'H': 25,   # VERSION
        'I': 12,   # DURACION
        'J': 12,   # HORA
        'K': 25,   # EMISORA/SITE
        'L': 30,   # PROGRAMA/TIPO DE SITE
        'M': 12,   # BREAK
        'N': 12,   # POS. SPOT
        'O': 15,   # INVERSION
        'P': 30,   # SECTOR
        'Q': 20,   # CATEGORIA
        'R': 15,   # ITEM
        'S': 20,   # AGENCIA
        'T': 25,   # ANUNCIANTE
        'U': 15,   # REGION/ÁMBITO
        'V': 12,   # ANCHO
        'W': 12,   # ALTO
        'X': 15,   # GENERO
        'Y': 25    # EDITORA
    }

    for col_letter, width in ANCHOS.items():
        try:
            worksheet.column_dimensions[col_letter].width = width
        except:
            pass


# ==========================================
# FUNCIONES DE FORMATO - OUTVIEW
# ==========================================

def aplicar_formato_outview(worksheet, df: pd.DataFrame):
    """
    Aplica formato a hoja OutView
    """

    # 1. Formato metadatos (fila 2)
    header_fill = PatternFill(
        start_color='D9E1F2',
        end_color='D9E1F2',
        fill_type='solid'
    )
    bold_font = Font(bold=True)

    for col in range(1, 3):  # Columnas A-B
        worksheet.cell(row=2, column=col).font = bold_font
        worksheet.cell(row=2, column=col).fill = header_fill

    # 2. Formato headers (fila 9)
    header_fill_datos = PatternFill(
        start_color='EC4899',  # Magenta
        end_color='EC4899',
        fill_type='solid'
    )
    header_font = Font(bold=True, color='FFFFFF')

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=9, column=col)
        cell.fill = header_fill_datos
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # 3. Formato fechas
    for row in range(10, len(df) + 10):
        worksheet.cell(row=row, column=1).number_format = 'DD/MM/YYYY'

    # 4. Formato Tarifa Real ($)
    cols_dict = {col: idx + 1 for idx, col in enumerate(df.columns)}
    if 'Tarifa Real ($)' in cols_dict:
        col_tarifa = cols_dict['Tarifa Real ($)']
        for row in range(10, len(df) + 10):
            worksheet.cell(row=row, column=col_tarifa).number_format = '#,##0.00'

    # 5. Formato coordenadas
    if 'Latitud' in cols_dict:
        col_lat = cols_dict['Latitud']
        for row in range(10, len(df) + 10):
            worksheet.cell(row=row, column=col_lat).number_format = '0.000000'

    if 'Longitud' in cols_dict:
        col_lon = cols_dict['Longitud']
        for row in range(10, len(df) + 10):
            worksheet.cell(row=row, column=col_lon).number_format = '0.000000'

    # 6. Ajustar anchos
    ajustar_anchos_outview(worksheet)


def ajustar_anchos_outview(worksheet):
    """
    Ajusta anchos de columnas para OutView
    """
    ANCHOS = {
        'A': 12,   # Fecha
        'B': 6,    # AÑO
        'C': 12,   # MES
        'D': 8,    # SEMANA
        'E': 15,   # Medio
        'F': 15,   # Marca
        'G': 20,   # Producto
        'H': 25,   # Versión
        'I': 18,   # Tipo Elemento
        'J': 18,   # Q versiones
        'K': 20,   # Distrito
        'L': 25,   # Avenida
        'M': 15,   # Nro Calle
        'N': 20,   # Orientación
        'O': 15,   # Tarifa Real
        'P': 30,   # Sector
        'Q': 20,   # Categoría
        'R': 15,   # Item
        'S': 20,   # Agencia
        'T': 25,   # Anunciante
        'U': 15,   # Región
        'V': 15,   # Latitud
        'W': 15,   # Longitud
        'X': 15,   # +1 Superficie
        'Y': 12,   # Conteo mensual
        'Z': 25    # Proveedor
    }

    for col_letter, width in ANCHOS.items():
        try:
            worksheet.column_dimensions[col_letter].width = width
        except:
            pass


# ==========================================
# FUNCIONES DE FORMATO - CONSOLIDADO
# ==========================================

def aplicar_formato_consolidado(worksheet, df: pd.DataFrame):
    """
    Aplica formato a hoja Consolidado
    """

    # 1. Formato metadatos (fila 2)
    header_fill = PatternFill(
        start_color='D9E1F2',
        end_color='D9E1F2',
        fill_type='solid'
    )
    bold_font = Font(bold=True)

    for col in range(1, 3):  # Columnas A-B
        worksheet.cell(row=2, column=col).font = bold_font
        worksheet.cell(row=2, column=col).fill = header_fill

    # 2. Formato headers (fila 9)
    header_fill_datos = PatternFill(
        start_color='8B5CF6',  # Púrpura (diferente de Monitor y OutView)
        end_color='8B5CF6',
        fill_type='solid'
    )
    header_font = Font(bold=True, color='FFFFFF')

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=9, column=col)
        cell.fill = header_fill_datos
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # 3. Formato fechas (columna A - FECHA)
    for row in range(10, len(df) + 10):
        worksheet.cell(row=row, column=1).number_format = 'DD/MM/YYYY'

    # 4. Formato INVERSIÓN REAL (columna P, #16)
    for row in range(10, len(df) + 10):
        worksheet.cell(row=row, column=16).number_format = '#,##0.00'

    # 5. Formato coordenadas (columnas W-X, #23-24)
    for row in range(10, len(df) + 10):
        # ANCHO / LATITUD
        worksheet.cell(row=row, column=23).number_format = '0.000000'
        # ALTO / LONGITUD
        worksheet.cell(row=row, column=24).number_format = '0.000000'

    # 6. Ajustar anchos
    ajustar_anchos_consolidado(worksheet)


def ajustar_anchos_consolidado(worksheet):
    """
    Ajusta anchos de columnas para Consolidado (27 columnas)
    """
    ANCHOS = {
        'A': 12,   # FECHA
        'B': 6,    # AÑO
        'C': 12,   # MES
        'D': 8,    # SEMANA
        'E': 15,   # MEDIO
        'F': 15,   # MARCA
        'G': 20,   # PRODUCTO
        'H': 25,   # VERSIÓN
        'I': 12,   # DURACIÓN
        'J': 18,   # TIPO ELEMENTO
        'K': 18,   # TIME / Q VERSIONES
        'L': 25,   # EMISORA / DISTRITO
        'M': 30,   # PROGRAMA / AVENIDA
        'N': 12,   # BREAK / CALLE
        'O': 20,   # POS. SPOT / ORIENTACIÓN
        'P': 15,   # INVERSIÓN REAL
        'Q': 30,   # SECTOR
        'R': 20,   # CATEGORÍA
        'S': 15,   # ÍTEM
        'T': 20,   # AGENCIA
        'U': 25,   # ANUNCIANTE
        'V': 15,   # REGIÓN
        'W': 18,   # ANCHO / LATITUD
        'X': 18,   # ALTO / LONGITUD
        'Y': 18,   # GEN / +1 SUPERFICIE
        'Z': 12,   # Q ELEMENTOS
        'AA': 25   # EDITORA / PROVEEDOR
    }

    for col_letter, width in ANCHOS.items():
        try:
            worksheet.column_dimensions[col_letter].width = width
        except:
            pass
