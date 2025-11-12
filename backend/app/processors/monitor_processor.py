# backend/app/processors/monitor_processor.py
"""
Procesador de archivos Monitor de Kantar Ibope Media

Este módulo procesa archivos .txt pipe-delimited de Kantar Ibope que contienen
datos de inversión publicitaria en medios ATL (TV, Cable, Radio, Revista, Diarios).

La transformación crítica es la factorización de inversión:
    INVERSION_NUEVA = INVERSION_ORIGINAL × FACTOR_MEDIO

Factores de conversión:
    - TV: 0.255
    - CABLE: 0.425
    - RADIO: 0.425
    - REVISTA: 0.14875
    - DIARIOS: 0.14875
"""

from __future__ import annotations

import io
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logger = logging.getLogger('mougli.monitor')

# Ruta al archivo de configuración de factores
CONFIG_PATH = Path(__file__).parent / 'factores_config.json'


class MonitorProcessor:
    """
    Procesador de archivos Monitor de Kantar Ibope Media

    Atributos:
        FACTORES (dict): Factores de conversión por medio
        metadatos (dict): Metadatos extraídos del archivo
        df (DataFrame): Datos procesados

    Métodos públicos:
        procesar(file_content: str) -> pd.DataFrame
        generar_excel() -> io.BytesIO
    """

    # Orden exacto de columnas en el output
    COLUMNAS_ORDENADAS = [
        'DIA',
        'AÑO',
        'MES',
        'SEMANA',
        'MEDIO',
        'MARCA',
        'PRODUCTO',
        'VERSION',
        'VERSION DESCRIPTIVA',
        'DURACION',
        'DUR.T.',
        'TIPO',
        'HORA',
        'EMISORA/SITE',
        'PROGRAMA/TIPO DE SITE',
        'BREAK',
        'POS. SPOT',
        'PAG/POSICION',
        'INVERSION',
        'TIPO TARIFA',
        'SECTOR',
        'CATEGORIA',
        'ITEM',
        'CALIDAD',
        'GENERO',
        'AGENCIA',
        'ANUNCIANTE',
        'SECCION/COMERC.',
        'BLOQUE/TOT.PAGS',
        'EDITORA',
        'EDICION',
        'COLOR',
        'SPOTS',
        'AREA',
        '%PAG.',
        'DES. POSICION',
        'ANCHO',
        'ALTO',
        'REGION/ÁMBITO',
        'CORTE LOCAL',
        'RUC'
    ]

    # Nombres de meses en español
    MESES = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }

    # Anchos de columna para Excel (en caracteres)
    ANCHOS_COLUMNA = {
        'A': 12,   # DIA
        'B': 6,    # AÑO
        'C': 12,   # MES
        'D': 8,    # SEMANA
        'E': 10,   # MEDIO
        'F': 15,   # MARCA
        'G': 15,   # PRODUCTO
        'H': 20,   # VERSION
        'I': 25,   # VERSION DESCRIPTIVA
        'J': 10,   # DURACION
        'K': 8,    # DUR.T.
        'L': 20,   # TIPO
        'M': 8,    # HORA
        'N': 20,   # EMISORA/SITE
        'O': 30,   # PROGRAMA/TIPO DE SITE
        'P': 8,    # BREAK
        'Q': 10,   # POS. SPOT
        'R': 12,   # PAG/POSICION
        'S': 12,   # INVERSION
        'T': 12,   # TIPO TARIFA
        'U': 30,   # SECTOR
        'V': 20,   # CATEGORIA
        'W': 15,   # ITEM
        'X': 10,   # CALIDAD
        'Y': 15,   # GENERO
        'Z': 20,   # AGENCIA
        'AA': 25,  # ANUNCIANTE
        'AB': 20,  # SECCION/COMERC.
        'AC': 15,  # BLOQUE/TOT.PAGS
        'AD': 25,  # EDITORA
        'AE': 10,  # EDICION
        'AF': 10,  # COLOR
        'AG': 8,   # SPOTS
        'AH': 12,  # AREA
        'AI': 8,   # %PAG.
        'AJ': 20,  # DES. POSICION
        'AK': 10,  # ANCHO
        'AL': 10,  # ALTO
        'AM': 15,  # REGION/ÁMBITO
        'AN': 12,  # CORTE LOCAL
        'AO': 15   # RUC
    }

    def __init__(self, factores_custom: Optional[Dict[str, float]] = None):
        """
        Inicializa el procesador de Monitor

        Args:
            factores_custom: Factores de conversión personalizados (opcional)
        """
        # Cargar factores desde config o usar custom
        if factores_custom:
            self.FACTORES = factores_custom
        else:
            self.FACTORES = self._cargar_factores()

        self.metadatos: Dict = {}
        self.df: Optional[pd.DataFrame] = None
        self.metadatos_originales: List[str] = []

    def _cargar_factores(self) -> Dict[str, float]:
        """Carga factores de conversión desde archivo JSON"""
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config['monitor']
        except FileNotFoundError:
            logger.warning(f"Archivo de configuración no encontrado: {CONFIG_PATH}. Usando factores por defecto.")
            return {
                'TV': 0.255,
                'CABLE': 0.425,
                'RADIO': 0.425,
                'REVISTA': 0.14875,
                'DIARIOS': 0.14875
            }
        except (json.JSONDecodeError, KeyError) as e:
            logger.error(f"Error leyendo configuración: {e}. Usando factores por defecto.")
            return {
                'TV': 0.255,
                'CABLE': 0.425,
                'RADIO': 0.425,
                'REVISTA': 0.14875,
                'DIARIOS': 0.14875
            }

    def procesar(self, file_content: str) -> pd.DataFrame:
        """
        Procesa el contenido del archivo Monitor

        Args:
            file_content: Contenido del archivo .txt como string

        Returns:
            DataFrame con datos procesados

        Raises:
            ValueError: Si el archivo es inválido
        """
        logger.info("Iniciando procesamiento de archivo Monitor")

        # 1. Validar archivo
        self._validar_archivo(file_content)

        # 2. Dividir en líneas
        lineas = file_content.split('\n')
        lineas = [l.strip() for l in lineas if l.strip()]  # Eliminar vacías

        logger.info(f"Archivo tiene {len(lineas)} líneas")

        # 3. Extraer metadatos originales (primeras 4 líneas)
        self.metadatos_originales = lineas[:4]

        # 4. Encontrar header
        header_idx = self._encontrar_header(lineas)
        logger.info(f"Header encontrado en línea {header_idx + 1}")

        # 5. Parsear header
        header_line = lineas[header_idx]
        columnas = [col.strip() for col in header_line.split('|')]

        # Renombrar primera columna '#' a 'ID'
        if columnas[0] == '#':
            columnas[0] = 'ID'

        logger.info(f"Encontradas {len(columnas)} columnas")

        # 6. Parsear datos
        datos_lineas = lineas[header_idx + 1:]
        datos = []

        for i, linea in enumerate(datos_lineas, start=1):
            valores = linea.split('|')

            # Validar número de columnas
            if len(valores) != len(columnas):
                logger.warning(
                    f"Línea {header_idx + 1 + i} tiene {len(valores)} valores, "
                    f"esperado {len(columnas)}. Skippeando."
                )
                continue

            datos.append(valores)

        logger.info(f"Parseadas {len(datos)} filas de datos")

        # 7. Crear DataFrame
        self.df = pd.DataFrame(datos, columns=columnas)

        # 8. Eliminar columna ID (ya no la necesitamos)
        if 'ID' in self.df.columns:
            self.df = self.df.drop(columns=['ID'])

        # 9. Limpiar datos
        self._limpiar_datos()

        # 10. Convertir SUPLEMENTO a DIARIOS
        self.df['MEDIO'] = self.df['MEDIO'].replace('SUPLEMENTO', 'DIARIOS')
        logger.info("Convertidos SUPLEMENTO → DIARIOS")

        # 11. Aplicar factores de conversión (CRÍTICO)
        self._aplicar_factores()

        # 12. Agregar columnas derivadas
        self._agregar_columnas_derivadas()

        # 13. Reordenar columnas
        self._reordenar_columnas()

        # 14. Calcular metadatos
        self._calcular_metadatos()

        logger.info(f"Procesamiento completado. DataFrame final: {len(self.df)} filas × {len(self.df.columns)} columnas")

        return self.df

    def _validar_archivo(self, file_content: str) -> None:
        """Valida estructura básica del archivo"""
        lineas = [l.strip() for l in file_content.split('\n') if l.strip()]

        # Mínimo 6 líneas (4 meta + 1 header + 1 dato)
        if len(lineas) < 6:
            raise ValueError("Archivo demasiado corto. Mínimo 6 líneas esperadas.")

        # Header debe contener #|MEDIO|
        header_encontrado = any('#|MEDIO|' in line for line in lineas[:10])
        if not header_encontrado:
            raise ValueError("Header #|MEDIO| no encontrado en primeras 10 líneas")

        # Al menos una línea de datos después del header
        header_idx = next(i for i, line in enumerate(lineas) if '#|MEDIO|' in line)
        if header_idx == len(lineas) - 1:
            raise ValueError("No hay datos después del header")

    def _encontrar_header(self, lineas: List[str]) -> int:
        """
        Encuentra el índice de la línea de header

        Returns:
            Índice de la línea que contiene #|MEDIO|
        """
        for i, linea in enumerate(lineas):
            if '#|MEDIO|' in linea:
                return i

        raise ValueError("No se encontró header #|MEDIO| en el archivo")

    def _limpiar_datos(self) -> None:
        """Limpia y convierte tipos de datos"""
        logger.info("Limpiando datos...")

        # 1. Convertir fecha DIA
        if 'DIA' in self.df.columns:
            self.df['DIA'] = pd.to_datetime(
                self.df['DIA'],
                format='%d/%m/%Y',
                dayfirst=True,
                errors='coerce'
            )

            fechas_invalidas = self.df['DIA'].isna().sum()
            if fechas_invalidas > 0:
                logger.warning(f"{fechas_invalidas} fechas inválidas convertidas a NaT")

        # 2. Convertir INVERSION a numérico
        if 'INVERSION' in self.df.columns:
            # Eliminar comas si las hay (formato europeo)
            self.df['INVERSION'] = self.df['INVERSION'].astype(str).str.replace(',', '')

            # Convertir a float
            self.df['INVERSION'] = pd.to_numeric(
                self.df['INVERSION'],
                errors='coerce'
            ).fillna(0)

            logger.info(f"INVERSION: min={self.df['INVERSION'].min():.2f}, max={self.df['INVERSION'].max():.2f}")

        # 3. Convertir otras columnas numéricas si existen
        columnas_numericas = ['AREA', '%PAG.', 'ANCHO', 'ALTO', 'SPOTS']

        for col in columnas_numericas:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(
                    self.df[col].astype(str).str.replace(',', ''),
                    errors='coerce'
                ).fillna(0)

        # 4. Limpiar campos de texto (strip whitespace)
        columnas_texto = ['MEDIO', 'MARCA', 'PRODUCTO', 'ANUNCIANTE', 'SECTOR', 'CATEGORIA', 'REGION/ÁMBITO']

        for col in columnas_texto:
            if col in self.df.columns:
                self.df[col] = self.df[col].astype(str).str.strip()

    def _aplicar_factores(self) -> None:
        """
        Aplica factores de conversión a la columna INVERSION

        CRÍTICO: Esta función SOBRESCRIBE la columna INVERSION original
        con los valores factorizados.

        Fórmula: INVERSION_NUEVA = INVERSION_ORIGINAL × FACTOR_MEDIO
        """
        logger.info("Aplicando factores de conversión...")

        if 'INVERSION' not in self.df.columns or 'MEDIO' not in self.df.columns:
            logger.warning("Columnas INVERSION o MEDIO no encontradas. Skippeando factorización.")
            return

        # Guardar valores originales para log
        inversion_original_total = self.df['INVERSION'].sum()

        # Aplicar factor según medio
        self.df['INVERSION'] = self.df.apply(
            lambda row: row['INVERSION'] * self.FACTORES.get(row['MEDIO'], 1.0),
            axis=1
        )

        # Redondear a 2 decimales
        self.df['INVERSION'] = self.df['INVERSION'].round(2)

        inversion_factorizada_total = self.df['INVERSION'].sum()

        logger.info(f"Factorización completada:")
        logger.info(f"  - Inversión original total: ${inversion_original_total:,.2f}")
        logger.info(f"  - Inversión factorizada total: ${inversion_factorizada_total:,.2f}")
        logger.info(f"  - Reducción: {(1 - inversion_factorizada_total/inversion_original_total)*100:.1f}%")

        # Logear medios sin factor
        medios_unicos = self.df['MEDIO'].unique()
        medios_sin_factor = [m for m in medios_unicos if m not in self.FACTORES]

        if medios_sin_factor:
            logger.warning(f"Medios sin factor de conversión (usando 1.0): {medios_sin_factor}")

    def _agregar_columnas_derivadas(self) -> None:
        """Agrega columnas derivadas: AÑO, MES, SEMANA"""
        logger.info("Agregando columnas derivadas...")

        if 'DIA' not in self.df.columns:
            logger.warning("Columna DIA no encontrada. No se pueden agregar columnas derivadas.")
            return

        # AÑO
        self.df['AÑO'] = self.df['DIA'].dt.year

        # MES (nombre en español)
        self.df['MES'] = self.df['DIA'].dt.month.map(self.MESES)

        # SEMANA (ISO week number)
        self.df['SEMANA'] = self.df['DIA'].dt.isocalendar().week

        logger.info("Columnas derivadas agregadas: AÑO, MES, SEMANA")

    def _reordenar_columnas(self) -> None:
        """Reordena columnas según el orden especificado"""
        logger.info("Reordenando columnas...")

        # Agregar columnas faltantes con valores vacíos
        columnas_faltantes = set(self.COLUMNAS_ORDENADAS) - set(self.df.columns)

        for col in columnas_faltantes:
            self.df[col] = ''
            logger.info(f"Columna '{col}' no encontrada, agregada como vacía")

        # Reordenar
        self.df = self.df[self.COLUMNAS_ORDENADAS]

        logger.info(f"Columnas reordenadas: {len(self.df.columns)} columnas en orden correcto")

    def _calcular_metadatos(self) -> None:
        """Calcula metadatos del DataFrame procesado"""
        if self.df is None or len(self.df) == 0:
            return

        self.metadatos = {
            'filas': len(self.df),
            'rango_fechas': f"{self.df['DIA'].min().strftime('%d/%m/%Y')} - {self.df['DIA'].max().strftime('%d/%m/%Y')}",
            'marcas_unicas': self.df['MARCA'].nunique(),
            'sectores': ', '.join(self.df['SECTOR'].unique()[:3].tolist()) + '...' if len(self.df['SECTOR'].unique()) > 3 else ', '.join(self.df['SECTOR'].unique()),
            'categorias': ', '.join(self.df['CATEGORIA'].unique()[:5].tolist()) + '...' if len(self.df['CATEGORIA'].unique()) > 5 else ', '.join(self.df['CATEGORIA'].unique()),
            'regiones': ', '.join(sorted(self.df['REGION/ÁMBITO'].unique()))
        }

        logger.info(f"Metadatos calculados: {self.metadatos}")

    def generar_excel(self) -> io.BytesIO:
        """
        Genera archivo Excel con metadatos y datos procesados

        Returns:
            BytesIO con archivo Excel

        Raises:
            ValueError: Si no hay datos procesados
        """
        if self.df is None or len(self.df) == 0:
            raise ValueError("No hay datos procesados para generar Excel")

        logger.info("Generando archivo Excel...")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Monitor"

        # 1. Escribir metadatos (filas 1-8)
        metadatos_df = self._crear_dataframe_metadatos()

        for r_idx, row in enumerate(dataframe_to_rows(metadatos_df, index=False, header=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Formato para metadatos
                if c_idx == 1:  # Columna "Descripción"
                    cell.font = Font(bold=True)

        # 2. Escribir headers (fila 9)
        header_row = 9
        for c_idx, col_name in enumerate(self.df.columns, start=1):
            cell = ws.cell(row=header_row, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # 3. Escribir datos (fila 10+)
        for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), start=10):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Formato para fechas
                if c_idx == 1 and isinstance(value, pd.Timestamp):
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='right')

                # Formato para números
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    cell.alignment = Alignment(horizontal='right')

        # 4. Ajustar anchos de columna
        self._ajustar_columnas(ws)

        # 5. Guardar en BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Excel generado exitosamente")

        return output

    def _crear_dataframe_metadatos(self) -> pd.DataFrame:
        """Crea DataFrame de metadatos para las filas 1-8 del Excel"""
        metadatos_data = [
            ['Descripción', 'Valor'],
            ['Filas', self.metadatos.get('filas', 0)],
            ['Rango de fechas', self.metadatos.get('rango_fechas', '')],
            ['Marcas / Anunciantes', self.metadatos.get('marcas_unicas', 0)],
            ['Sectores', self.metadatos.get('sectores', '')],
            ['Categorías', self.metadatos.get('categorias', '')],
            ['Regiones', self.metadatos.get('regiones', '')],
            ['', '']  # Fila vacía
        ]

        return pd.DataFrame(metadatos_data)

    def _ajustar_columnas(self, worksheet) -> None:
        """Ajusta anchos de columnas en worksheet"""
        for col_letter, width in self.ANCHOS_COLUMNA.items():
            worksheet.column_dimensions[col_letter].width = width

    def validar_datos_procesados(self) -> List[str]:
        """
        Valida calidad de datos procesados

        Returns:
            Lista de warnings
        """
        if self.df is None:
            return ["No hay datos procesados"]

        warnings = []

        # 1. Fechas válidas (al menos 90%)
        fechas_validas = self.df['DIA'].notna().sum()
        porcentaje = (fechas_validas / len(self.df)) * 100
        if porcentaje < 90:
            warnings.append(f"Solo {porcentaje:.1f}% fechas válidas (esperado >90%)")

        # 2. Inversión: al menos 20% con valores > 0
        inv_positivas = (self.df['INVERSION'] > 0).sum()
        porcentaje_inv = (inv_positivas / len(self.df)) * 100
        if porcentaje_inv < 20:
            warnings.append(f"Solo {porcentaje_inv:.1f}% inversiones >0 (¿correcto?)")

        # 3. Columnas críticas no vacías
        columnas_criticas = ['MEDIO', 'MARCA', 'ANUNCIANTE']
        for col in columnas_criticas:
            if col in self.df.columns:
                vacias = self.df[col].isna().sum()
                if vacias > 0:
                    warnings.append(f"Columna '{col}' tiene {vacias} valores vacíos")

        # 4. Medios válidos
        medios_esperados = {'TV', 'CABLE', 'RADIO', 'REVISTA', 'DIARIOS'}
        medios_encontrados = set(self.df['MEDIO'].unique())
        medios_raros = medios_encontrados - medios_esperados
        if medios_raros:
            warnings.append(f"Medios no esperados: {medios_raros}")

        return warnings


def procesar_monitor_txt(
    file_content: str,
    factores_custom: Optional[Dict[str, float]] = None
) -> io.BytesIO:
    """
    Función de conveniencia para procesar archivo Monitor y retornar Excel

    Args:
        file_content: Contenido del archivo .txt como string
        factores_custom: Factores de conversión personalizados (opcional)

    Returns:
        BytesIO con archivo Excel

    Raises:
        ValueError: Si archivo inválido
    """
    processor = MonitorProcessor(factores_custom=factores_custom)
    processor.procesar(file_content)

    # Validar datos procesados
    warnings = processor.validar_datos_procesados()
    if warnings:
        for warning in warnings:
            logger.warning(f"Validación: {warning}")

    return processor.generar_excel()
