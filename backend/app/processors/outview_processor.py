# backend/app/processors/outview_processor.py
"""
Procesador de archivos OutView de Kantar Ibope Media

Este m√≥dulo procesa archivos Excel (.xlsx) de Kantar Ibope que contienen
datos de inversi√≥n publicitaria en medios OOH (Out of Home - V√≠a P√∫blica).

La transformaci√≥n cr√≠tica es un c√°lculo complejo de tarifa real en 9 pasos:
    1. Tarifa_USD = Tarifa S/. √∑ 3
    2. Denominador_1 = Count(por Fecha + ubicaci√≥n)
    3. Tarifa_1 = Tarifa_USD √∑ Denominador_1
    4. Denominador_2 = Count(por Mes + ubicaci√≥n)
    5. Tarifa_2 = Sum(Tarifa_1 por Mes)
    6. Tarifa_3 = Min(Tarifa_2, Tope por Tipo Elemento)
    7. Tarifa_4 = Tarifa_3 √∑ Denominador_2
    8. Factor = 0.4 (LED) o 0.8 (otros)
    9. Tarifa Real ($) = Tarifa_4 √ó Factor

Caracter√≠stica √∫nica: Todas las filas tienen Tarifa > 0 (solo Lima, sin provincias).
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logger = logging.getLogger('mougli.outview')


class OutViewProcessor:
    """
    Procesador de archivos OutView de Kantar Ibope Media

    Atributos:
        TOPES_TARIFA (dict): Topes m√°ximos por tipo de elemento
        TIPO_CAMBIO_USD (float): Tipo de cambio S/. a USD
        FACTOR_LED (float): Factor para PANTALLA LED
        FACTOR_OTROS (float): Factor para otros elementos
        df (DataFrame): Datos procesados

    M√©todos p√∫blicos:
        procesar(file_content: bytes) -> pd.DataFrame
        generar_excel() -> io.BytesIO
    """

    # Topes de tarifa por tipo de elemento (en USD)
    TOPES_TARIFA = {
        'BANDEROLA': 16000.00,
        'CLIP': 800.00,
        'MINIPOLAR': 1333.33,
        'PALETA': 800.00,
        'PANEL': 2433.33,
        'PANEL CARRETERO': 6666.67,
        'PANTALLA LED': 7200.00,
        'PARADERO': 1066.67,
        'PRISMA': 3733.33,
        'QUIOSCO': 800.00,
        'RELOJ': 1120.00,
        'TORRE UNIPOLAR': 4000.00,
        'TOTEM': 1266.67,
        'VALLA': 800.00,
        'VALLA ALTA': 1733.33
    }

    # Constantes de conversi√≥n
    TIPO_CAMBIO_USD = 3.0
    FACTOR_LED = 0.4
    FACTOR_OTROS = 0.8

    # Nombres de meses en espa√±ol
    MESES = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }

    # Orden exacto de columnas output (33 columnas)
    COLUMNAS_OUTPUT = [
        'Fecha',
        'A√ëO',
        'MES',
        'SEMANA',
        'NombreBase',
        'Medio',
        'Proveedor',
        'Cod.Proveedor',
        'Tipo Elemento',
        'Distrito',
        'Avenida',
        'Nro Calle/Cuadra',
        'Orientaci√≥n de V√≠a',
        'Sector',
        'Categor√≠a',
        'Item',
        'Marca',
        'Producto',
        'Versi√≥n',
        'Agencia',
        'Anunciante',
        'Regi√≥n',
        'Tipo Tarifa',
        'Duraci√≥n (Seg)',
        'Latitud',
        'Longitud',
        'EstadoAviso',
        'RUC',
        'Q versiones por elemento Mes',
        '+1 Superficie',
        'Tarifa √ó Superficie (1ra por C√≥digo √∫nico)',
        'Conteo mensual',
        'Tarifa Real ($)'
    ]

    # Anchos de columna para Excel
    ANCHOS_COLUMNA = {
        'A': 12,   # Fecha
        'B': 6,    # A√ëO
        'C': 12,   # MES
        'D': 8,    # SEMANA
        'E': 18,   # NombreBase
        'F': 12,   # Medio
        'G': 20,   # Proveedor
        'H': 15,   # Cod.Proveedor
        'I': 18,   # Tipo Elemento
        'J': 20,   # Distrito
        'K': 25,   # Avenida
        'L': 15,   # Nro Calle/Cuadra
        'M': 20,   # Orientaci√≥n de V√≠a
        'N': 30,   # Sector
        'O': 20,   # Categor√≠a
        'P': 15,   # Item
        'Q': 15,   # Marca
        'R': 20,   # Producto
        'S': 30,   # Versi√≥n
        'T': 20,   # Agencia
        'U': 25,   # Anunciante
        'V': 10,   # Regi√≥n
        'W': 12,   # Tipo Tarifa
        'X': 12,   # Duraci√≥n (Seg)
        'Y': 12,   # Latitud
        'Z': 12,   # Longitud
        'AA': 12,  # EstadoAviso
        'AB': 15,  # RUC
        'AC': 25,  # Q versiones por elemento Mes
        'AD': 15,  # +1 Superficie
        'AE': 35,  # Tarifa √ó Superficie
        'AF': 15,  # Conteo mensual
        'AG': 15   # Tarifa Real ($)
    }

    def __init__(self):
        """Inicializa el procesador de OutView"""
        self.df: Optional[pd.DataFrame] = None
        self.metadatos: Dict = {}

    def procesar(self, file_content: bytes) -> pd.DataFrame:
        """
        Procesa el contenido del archivo OutView (17 pasos)

        Args:
            file_content: Contenido del archivo Excel

        Returns:
            DataFrame con datos procesados

        Raises:
            ValueError: Si el archivo es inv√°lido
        """
        try:
            logger.info("=" * 60)
            logger.info("üöÄ INICIANDO PROCESAMIENTO DE ARCHIVO OUTVIEW")
            logger.info("=" * 60)

            # PASO 1: Leer Excel (CR√çTICO: skiprows=1 porque fila 1 est√° vac√≠a)
            logger.info("PASO 1: Leer Excel")
            self.df = self._leer_excel(file_content)
            logger.info(f"‚úÖ OutView le√≠do: {len(self.df)} filas, {len(self.df.columns)} columnas")

            # PASO 2-3: Fechas derivadas
            logger.info("PASO 2-3: Procesar fechas")
            self.df = self._procesar_fechas(self.df)
            self.df = self._extraer_mes_nombrebase(self.df)

            # PASO 4: Identificadores √∫nicos
            logger.info("PASO 4: Crear identificadores √∫nicos")
            self.df = self._crear_codigo_unico(self.df)
            self.df = self._crear_codigo_pieza(self.df)

            # PASO 5-6: Denominadores
            logger.info("PASO 5-6: Calcular denominadores")
            self.df = self._calcular_denominador_1(self.df)
            self.df = self._calcular_denominador_2(self.df)

            # PASO 7-15: Tarifas (9 pasos)
            logger.info("PASO 7-15: Calcular tarifas")
            self.df = self._calcular_tarifas(self.df)

            # PASO 16-19: Columnas finales
            logger.info("PASO 16-19: Calcular columnas finales")
            self.df = self._calcular_columnas_finales(self.df)

            # PASO 20: Reordenar columnas
            logger.info("PASO 20: Reordenar columnas")
            self.df = self._reordenar_columnas(self.df)

            # PASO 21: Calcular metadatos
            logger.info("PASO 21: Calcular metadatos")
            self._calcular_metadatos()

            logger.info("=" * 60)
            logger.info(f"‚úÖ PROCESAMIENTO COMPLETADO: {len(self.df)} filas √ó {len(self.df.columns)} columnas")
            logger.info("=" * 60)

            return self.df

        except ValueError as ve:
            logger.error(f"‚ùå Error de validaci√≥n: {str(ve)}")
            raise
        except Exception as e:
            logger.error(f"‚ùå Error inesperado en procesamiento: {type(e).__name__}: {str(e)}", exc_info=True)
            raise ValueError(f"Error procesando archivo OutView: {str(e)}")

    def _leer_excel(self, file_content: bytes) -> pd.DataFrame:
        """
        Lee archivo Excel saltando fila 1 vac√≠a

        CR√çTICO: La fila 1 SIEMPRE est√° vac√≠a en archivos OutView
        """
        try:
            logger.info(f"üìÑ Intentando leer Excel, tama√±o: {len(file_content)} bytes ({len(file_content) / 1024:.2f} KB)")

            # Intentar leer con skiprows=1 primero (formato est√°ndar)
            try:
                df = pd.read_excel(
                    io.BytesIO(file_content),
                    engine='openpyxl',  # Especificar engine expl√≠citamente
                    skiprows=1  # ‚ö†Ô∏è CR√çTICO: Salta fila 1 vac√≠a
                )
                logger.info(f"‚úÖ Excel le√≠do con skiprows=1: {len(df)} filas, {len(df.columns)} columnas")
            except Exception as e1:
                logger.warning(f"‚ö†Ô∏è No se pudo leer con skiprows=1: {e1}")
                logger.info("üîÑ Intentando sin skiprows...")

                # Intentar sin skiprows (algunos archivos pueden no tener fila vac√≠a)
                df = pd.read_excel(
                    io.BytesIO(file_content),
                    engine='openpyxl',
                    skiprows=0
                )
                logger.info(f"‚úÖ Excel le√≠do sin skiprows: {len(df)} filas, {len(df.columns)} columnas")

            logger.info(f"üìã Columnas encontradas ({len(df.columns)}): {list(df.columns)[:15]}")  # Primeras 15

            # Validar que no est√© vac√≠o
            if len(df) == 0:
                logger.error("‚ùå El archivo est√° vac√≠o (0 filas)")
                raise ValueError("El archivo Excel no contiene datos")

            # Validar columnas m√≠nimas requeridas
            columnas_requeridas = ['Fecha', 'NombreBase', 'Tarifa S/.', 'Tipo Elemento']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]

            if columnas_faltantes:
                logger.error(f"‚ùå Columnas faltantes: {columnas_faltantes}")
                logger.error(f"üìã Columnas disponibles: {list(df.columns)}")
                raise ValueError(f"Archivo OutView inv√°lido. Columnas faltantes: {', '.join(columnas_faltantes)}")

            logger.info("‚úÖ Validaci√≥n de columnas pasada")
            return df

        except ValueError as ve:
            raise
        except Exception as e:
            logger.error(f"‚ùå Error leyendo Excel: {type(e).__name__}: {str(e)}", exc_info=True)
            raise ValueError(f"Error leyendo archivo Excel: {str(e)}")

    def _procesar_fechas(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Extrae fechas derivadas: A√ëO, MES, SEMANA
        """
        try:
            logger.info("üîÑ Procesando fechas...")

            # Convertir Fecha a datetime
            df['Fecha'] = pd.to_datetime(
                df['Fecha'],
                format='%d/%m/%Y',
                dayfirst=True,
                errors='coerce'
            )

            fechas_invalidas = df['Fecha'].isna().sum()
            if fechas_invalidas > 0:
                logger.warning(f"‚ö†Ô∏è {fechas_invalidas} fechas inv√°lidas (NaT)")

            # A√ëO
            df['A√ëO'] = df['Fecha'].dt.year

            # SEMANA (ISO)
            df['SEMANA'] = df['Fecha'].dt.isocalendar().week

            # MES (texto espa√±ol)
            df['Mes_Codigo'] = df['Fecha'].dt.month
            df['MES'] = df['Mes_Codigo'].map(self.MESES)

            logger.info("‚úÖ Fechas derivadas agregadas: A√ëO, MES, SEMANA")
            return df

        except Exception as e:
            logger.error(f"‚ùå Error procesando fechas: {type(e).__name__}: {str(e)}", exc_info=True)
            raise ValueError(f"Error procesando fechas: {str(e)}")

    def _extraer_mes_nombrebase(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Extrae Mes del NombreBase para agrupaciones

        Formato NombreBase: OPW09MAR2023
        Extracci√≥n: posiciones [5:12] = "MAR2023"
        """
        df['Mes'] = df['NombreBase'].str[5:12]

        logger.info("Mes extra√≠do de NombreBase")

        return df

    def _crear_codigo_unico(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Crea Codigo_Unico (K_UNICO) - 13 campos

        Identifica un elemento f√≠sico √∫nico en una ubicaci√≥n durante un mes
        """
        df['Codigo_Unico'] = (
            df['Mes'].astype(str) + '|' +
            df['A√ëO'].astype(str) + '|' +
            df['Latitud'].astype(str) + '|' +
            df['Longitud'].astype(str) + '|' +
            df['Avenida'].astype(str) + '|' +
            df['Nro Calle/Cuadra'].astype(str) + '|' +
            df['Marca'].astype(str) + '|' +
            df['Tipo Elemento'].astype(str) + '|' +
            df['Orientaci√≥n de V√≠a'].astype(str) + '|' +
            df['Tarifa S/.'].astype(str) + '|' +
            df['Proveedor'].astype(str) + '|' +
            df['Distrito'].astype(str) + '|' +
            df['Cod.Proveedor'].astype(str)
        )

        logger.info(f"Codigo_Unico creado: {df['Codigo_Unico'].nunique()} c√≥digos √∫nicos")

        return df

    def _crear_codigo_pieza(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Crea Codigo_Pieza (K_PIEZA) - 16 campos

        Identifica una pieza publicitaria espec√≠fica (elemento + versi√≥n + semana)
        """
        df['Codigo_Pieza'] = (
            df['NombreBase'].astype(str) + '|' +
            df['Proveedor'].astype(str) + '|' +
            df['Tipo Elemento'].astype(str) + '|' +
            df['Distrito'].astype(str) + '|' +
            df['Orientaci√≥n de V√≠a'].astype(str) + '|' +
            df['Nro Calle/Cuadra'].astype(str) + '|' +
            df['Item'].astype(str) + '|' +
            df['Versi√≥n'].astype(str) + '|' +
            df['Latitud'].astype(str) + '|' +
            df['Longitud'].astype(str) + '|' +
            df['Categor√≠a'].astype(str) + '|' +
            df['Tarifa S/.'].astype(str) + '|' +
            df['Anunciante'].astype(str) + '|' +
            df['Mes'].astype(str) + '|' +
            df['A√ëO'].astype(str) + '|' +
            df['SEMANA'].astype(str)
        )

        logger.info(f"Codigo_Pieza creado: {df['Codigo_Pieza'].nunique()} piezas √∫nicas")

        return df

    def _calcular_denominador_1(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calcula Denominador_1: Contador por FECHA (d√≠a)

        Agrupa por fecha + ubicaci√≥n para contar apariciones diarias
        """
        denominador1 = df.groupby([
            'Fecha',  # ‚ö†Ô∏è Por D√çA
            'Proveedor',
            'Tipo Elemento',
            'Distrito',
            'Avenida',
            'Nro Calle/Cuadra',
            'Orientaci√≥n de V√≠a',
            'Marca'
        ]).transform('size')

        df['Denominador_1'] = denominador1

        logger.info(f"Denominador_1: min={df['Denominador_1'].min()}, max={df['Denominador_1'].max()}, avg={df['Denominador_1'].mean():.2f}")

        return df

    def _calcular_denominador_2(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calcula Denominador_2: Contador por MES

        Agrupa por mes + ubicaci√≥n para contar apariciones mensuales
        """
        denominador2 = df.groupby([
            'Mes',  # ‚ö†Ô∏è Por MES (no Fecha)
            'Proveedor',
            'Tipo Elemento',
            'Distrito',
            'Avenida',
            'Nro Calle/Cuadra',
            'Orientaci√≥n de V√≠a',
            'Marca'
        ]).transform('size')

        df['Denominador_2'] = denominador2

        logger.info(f"Denominador_2: min={df['Denominador_2'].min()}, max={df['Denominador_2'].max()}, avg={df['Denominador_2'].mean():.2f}")

        return df

    def _calcular_tarifas(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calcula tarifas en 9 pasos (PASO 5-12 + PASO 10 final)

        Este es el coraz√≥n del procesamiento OutView
        """
        logger.info("Iniciando c√°lculo de tarifas (9 pasos)...")

        # PASO 5: Tarifa_USD
        df['Tarifa_USD'] = df['Tarifa S/.'] / self.TIPO_CAMBIO_USD
        logger.info(f"Paso 5 - Tarifa_USD: ${df['Tarifa_USD'].sum():,.2f} total")

        # PASO 6: Tarifa_1 (prorratea por d√≠a)
        df['Tarifa_1'] = df.apply(
            lambda row: row['Tarifa_USD'] / row['Denominador_1']
                        if row['Denominador_1'] > 0
                        else 0,
            axis=1
        )
        logger.info(f"Paso 6 - Tarifa_1: ${df['Tarifa_1'].sum():,.2f} total")

        # PASO 7: Tarifa_2 (suma por mes)
        tarifa2 = df.groupby([
            'Mes',
            'Proveedor',
            'Tipo Elemento',
            'Distrito',
            'Avenida',
            'Nro Calle/Cuadra',
            'Orientaci√≥n de V√≠a',
            'Marca'
        ])['Tarifa_1'].transform('sum')
        df['Tarifa_2'] = tarifa2
        logger.info(f"Paso 7 - Tarifa_2: ${df['Tarifa_2'].sum():,.2f} total")

        # PASO 8: Tarifa_3 (aplicar tope)
        def aplicar_tope(row):
            tipo_elem = row['Tipo Elemento']
            tarifa2 = row['Tarifa_2']
            tope = self.TOPES_TARIFA.get(tipo_elem, float('inf'))

            if tope == float('inf'):
                logger.warning(f"Tipo desconocido sin tope: '{tipo_elem}'")

            return min(tarifa2, tope)

        df['Tarifa_3'] = df.apply(aplicar_tope, axis=1)
        logger.info(f"Paso 8 - Tarifa_3 (con topes): ${df['Tarifa_3'].sum():,.2f} total")

        # PASO 9: Tarifa_4 (prorratea por mes)
        df['Tarifa_4'] = df.apply(
            lambda row: row['Tarifa_3'] / row['Denominador_2']
                        if row['Denominador_2'] > 0
                        else 0,
            axis=1
        )
        logger.info(f"Paso 9 - Tarifa_4: ${df['Tarifa_4'].sum():,.2f} total")

        # PASO 10: Tarifa Real ($) - FINAL
        def aplicar_factor(row):
            if row['Tipo Elemento'] == 'PANTALLA LED':
                return row['Tarifa_4'] * self.FACTOR_LED
            else:
                return row['Tarifa_4'] * self.FACTOR_OTROS

        df['Tarifa Real ($)'] = df.apply(aplicar_factor, axis=1)

        # Redondear a 2 decimales
        df['Tarifa Real ($)'] = df['Tarifa Real ($)'].round(2)

        logger.info(f"Paso 10 - Tarifa Real ($): ${df['Tarifa Real ($)'].sum():,.2f} total")
        logger.info("C√°lculo de tarifas completado (9 pasos)")

        return df

    def _calcular_columnas_finales(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calcula columnas finales (PASO 16-19)
        """
        logger.info("Calculando columnas finales...")

        # PASO 16: Q versiones por elemento Mes
        versiones = df.groupby('Codigo_Unico')['Versi√≥n'].nunique()
        df['Q versiones por elemento Mes'] = df['Codigo_Unico'].map(versiones)
        logger.info(f"Q versiones: avg={df['Q versiones por elemento Mes'].mean():.2f}")

        # PASO 17: +1 Superficie
        superficie = df.groupby('Codigo_Pieza').transform('size')
        df['+1 Superficie'] = superficie
        logger.info(f"+1 Superficie: avg={df['+1 Superficie'].mean():.2f}")

        # PASO 18: Tarifa √ó Superficie
        df['Tarifa √ó Superficie (1ra por C√≥digo √∫nico)'] = (
            df['Tarifa Real ($)'] * df['+1 Superficie']
        ).round(2)
        logger.info(f"Tarifa √ó Superficie: ${df['Tarifa √ó Superficie (1ra por C√≥digo √∫nico)'].sum():,.2f} total")

        # PASO 19: Conteo mensual
        df['Conteo mensual'] = (~df.duplicated(subset='Codigo_Unico')).astype(int)
        elementos_unicos = df['Conteo mensual'].sum()
        logger.info(f"Conteo mensual: {elementos_unicos} elementos √∫nicos")

        return df

    def _reordenar_columnas(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reordena columnas seg√∫n especificaci√≥n exacta (PASO 20)
        """
        logger.info("Reordenando columnas...")

        # Agregar columnas faltantes como vac√≠as
        for col in self.COLUMNAS_OUTPUT:
            if col not in df.columns:
                df[col] = ''
                logger.info(f"Columna '{col}' agregada como vac√≠a")

        # Eliminar columnas temporales
        columnas_temp = [
            'Mes_Codigo', 'Codigo_Unico', 'Codigo_Pieza',
            'Denominador_1', 'Denominador_2', 'Tarifa_USD',
            'Tarifa_1', 'Tarifa_2', 'Tarifa_3', 'Tarifa_4', 'Mes'
        ]

        for col in columnas_temp:
            if col in df.columns:
                df = df.drop(columns=[col])

        # Reordenar
        df = df[self.COLUMNAS_OUTPUT]

        logger.info(f"Columnas reordenadas: {len(df.columns)} columnas en orden correcto")

        return df

    def _calcular_metadatos(self) -> None:
        """Calcula metadatos del DataFrame procesado"""
        if self.df is None or len(self.df) == 0:
            return

        self.metadatos = {
            'filas': len(self.df),
            'rango_fechas': f"{self.df['Fecha'].min().strftime('%d/%m/%Y')} - {self.df['Fecha'].max().strftime('%d/%m/%Y')}",
            'marcas_unicas': self.df['Marca'].nunique(),
            'tipos': ', '.join(sorted(self.df['Tipo Elemento'].unique())[:5]),
            'proveedores': ', '.join(sorted(self.df['Proveedor'].unique())),
            'regiones': 'LIMA'  # Siempre LIMA en OutView
        }

        if self.df['Tipo Elemento'].nunique() > 5:
            self.metadatos['tipos'] += '...'

        logger.info(f"Metadatos calculados: {self.metadatos}")

    def generar_excel(self) -> io.BytesIO:
        """
        Genera archivo Excel con metadatos y datos procesados

        Returns:
            BytesIO con archivo Excel

        Raises:
            ValueError: Si no hay datos procesados
        """
        if self.df is None or len(self.df) == 0:
            raise ValueError("No hay datos procesados para generar Excel")

        logger.info(f"üìä Generando archivo Excel con {len(self.df)} filas...")

        try:
            # Crear workbook (NO usar write_only aqu√≠ porque necesitamos formatear)
            wb = Workbook()
            ws = wb.active
            ws.title = "OutView"
            logger.info("‚úÖ Workbook creado")

            # 1. Fila 1: VAC√çA (para mantener compatibilidad)
            ws.cell(row=1, column=1, value='')

            # 2. Escribir metadatos (filas 2-8)
            metadatos_df = self._crear_dataframe_metadatos()

            for r_idx, row in enumerate(dataframe_to_rows(metadatos_df, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Formato para metadatos
                    if c_idx == 1 and r_idx > 2:  # Columna "Descripci√≥n" (excepto header)
                        cell.font = Font(bold=True)

            # 3. Escribir headers (fila 9)
            header_row = 9
            for c_idx, col_name in enumerate(self.df.columns, start=1):
                cell = ws.cell(row=header_row, column=c_idx, value=col_name)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # 4. Escribir datos (fila 10+)
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), start=10):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Formato para fechas
                    if c_idx == 1 and isinstance(value, pd.Timestamp):
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal='right')

                    # Formato para n√∫meros con decimales
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        cell.alignment = Alignment(horizontal='right')

            # 5. Ajustar anchos de columna
            self._ajustar_columnas(ws)

            # 6. Guardar en BytesIO
            logger.info("üíæ Guardando Excel en memoria...")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_size_kb = len(output.getvalue()) / 1024
            logger.info(f"‚úÖ Excel generado exitosamente: {excel_size_kb:.2f} KB")

            return output

        except Exception as e:
            logger.error(f"‚ùå Error generando Excel: {type(e).__name__}: {str(e)}", exc_info=True)
            raise ValueError(f"Error generando archivo Excel: {str(e)}")

    def _crear_dataframe_metadatos(self) -> pd.DataFrame:
        """Crea DataFrame de metadatos para las filas 2-8 del Excel"""
        metadatos_data = [
            ['Descripci√≥n', 'Valor'],
            ['Filas', self.metadatos.get('filas', 0)],
            ['Rango de fechas', self.metadatos.get('rango_fechas', '')],
            ['Marcas / Anunciantes', self.metadatos.get('marcas_unicas', 0)],
            ['Tipo', self.metadatos.get('tipos', '')],
            ['Proveedor', self.metadatos.get('proveedores', '')],
            ['Regiones', self.metadatos.get('regiones', '')]
        ]

        return pd.DataFrame(metadatos_data)

    def _ajustar_columnas(self, worksheet) -> None:
        """Ajusta anchos de columnas en worksheet"""
        for col_letter, width in self.ANCHOS_COLUMNA.items():
            worksheet.column_dimensions[col_letter].width = width


def procesar_outview_excel(
    file_content: bytes
) -> io.BytesIO:
    """
    Funci√≥n de conveniencia para procesar archivo OutView y retornar Excel

    Args:
        file_content: Contenido del archivo Excel

    Returns:
        BytesIO con archivo Excel procesado

    Raises:
        ValueError: Si archivo inv√°lido
    """
    processor = OutViewProcessor()
    processor.procesar(file_content)

    return processor.generar_excel()
